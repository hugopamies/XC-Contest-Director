from openpyxl import Workbook
from openpyxl.styles import Font
from utils.storage import load_results
from scoring.scoring_engine import total_score
import json

def export_all_data_to_excel(filename="uav_results.xlsx"):
    with open("data/teams.json", "r", encoding="utf-8") as f:
        teams = json.load(f)
    results = load_results()

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "Summary"

    for category in ["academic", "clubs"]:
        ws = wb.create_sheet(title=category.capitalize())
        headers = [
            "Team ID", "Team Name", "Round", "Csol", "Cdes", "Tcircuit", "Tglide", "A60s",
            "Tcarga", "Takeoff (m)", "Internal Pilot", "Legal", "Landing OK",
            "Replacements", "Score"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h).font = Font(bold=True)

        row = 2
        for team in teams[category]:
            tid = str(team["id"])
            name = team["name"]
            rounds = results.get(category, {}).get(tid, [])

            for idx, score_entry in enumerate(rounds):
                inputs = score_entry.get("inputs", {}) if isinstance(score_entry, dict) else {}
                score = score_entry.get("score", 0) if isinstance(score_entry, dict) else score_entry

                ws.cell(row=row, column=1, value=int(tid))
                ws.cell(row=row, column=2, value=name)
                ws.cell(row=row, column=3, value=idx + 1)

                ws.cell(row=row, column=4, value=inputs.get("Csol"))
                ws.cell(row=row, column=5, value=inputs.get("Cdes"))
                ws.cell(row=row, column=6, value=inputs.get("Tcircuit"))
                ws.cell(row=row, column=7, value=inputs.get("Tglide"))
                ws.cell(row=row, column=8, value=inputs.get("A60s"))
                ws.cell(row=row, column=9, value=inputs.get("Tcarga"))
                ws.cell(row=row, column=10, value=inputs.get("takeoff_distance"))
                ws.cell(row=row, column=11, value=inputs.get("internal_pilot"))
                ws.cell(row=row, column=12, value=inputs.get("legal_flight"))
                ws.cell(row=row, column=13, value=inputs.get("good_landing"))
                ws.cell(row=row, column=14, value=inputs.get("used_replacement_parts"))
                ws.cell(row=row, column=15, value=score)
                row += 1

    wb.remove(ws_index)  # Remove empty "Summary" sheet
    wb.save(filename)
