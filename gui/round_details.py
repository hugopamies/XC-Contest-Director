from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QTabWidget, QTableWidget,
    QTableWidgetItem, QPushButton, QComboBox, QFileDialog, QMessageBox, QHBoxLayout
)
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtCore import Qt
import json
import math
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils.storage import load_results
from scoring.scoring_engine import compute_round_score, get_best_values_per_round
from utils.input_data_exporter import export_input_data_to_xlsx


class NumericTableWidgetItem(QTableWidgetItem):
    def __init__(self, text, numeric_value):
        super().__init__(text)
        self.numeric_value = numeric_value

    def __lt__(self, other):
        if isinstance(other, NumericTableWidgetItem):
            return self.numeric_value < other.numeric_value
        return super().__lt__(other)


class RoundDetailsTab(QWidget):

    def __init__(self):
        super().__init__()

        self.layout = QVBoxLayout(self)

        self.category_dropdown = QComboBox()
        self.category_dropdown.addItems(["Academic", "Clubs"])
        self.category_dropdown.currentTextChanged.connect(self.refresh_data)
        self.layout.addWidget(QLabel("Category"))
        self.layout.addWidget(self.category_dropdown)

        self.round_tabs = QTabWidget()
        self.layout.addWidget(self.round_tabs)

        button_row = QHBoxLayout()
        self.refresh_button = QPushButton("🔄 Refresh")
        self.refresh_button.clicked.connect(self.refresh_data)
        button_row.addWidget(self.refresh_button)

        self.export_pdf_btn = QPushButton("📄 Export to PDF")
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        button_row.addWidget(self.export_pdf_btn)

        self.export_excel_btn = QPushButton("📊 Export to Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        button_row.addWidget(self.export_excel_btn)

        self.export_inputs_btn = QPushButton("📥 Export Inputs to XLSX")
        self.export_inputs_btn.clicked.connect(self.export_inputs_to_xlsx)
        button_row.addWidget(self.export_inputs_btn)

        button_row.addStretch()
        self.layout.addLayout(button_row)

        self.setLayout(self.layout)

        self.teams = []
        self.results = {}
        self.refresh_data()

    def export_inputs_to_xlsx(self):
        if not self.results:
            QMessageBox.warning(self, "Warning", "No results data to export.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Input Data to XLSX",
            "XC_input_data.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filename:
            return

        try:
            export_input_data_to_xlsx(self.results, filename)
            QMessageBox.information(self, "Success", "Input data exported successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export input data:\n{e}")

    def refresh_data(self):
        category = self.category_dropdown.currentText()
        with open("data/teams.json", "r", encoding="utf-8") as f:
            teams_data = json.load(f)
            self.teams = teams_data.get(category, [])

        self.results = load_results()
        self.build_round_tabs(category)

    def build_round_tabs(self, category):
        self.round_tabs.clear()

        max_rounds = 0
        for team in self.teams:
            tid = str(team["id"])
            team_rounds = self.results.get(category, {}).get(tid, [])
            max_rounds = max(max_rounds, len(team_rounds))

        for round_index in range(max_rounds):
            tab = QWidget()
            layout = QVBoxLayout(tab)
            table = QTableWidget()
            layout.addWidget(table)
            tab.setLayout(layout)
            self.round_tabs.addTab(tab, f"Round {round_index + 1}")

            self.populate_table(table, category, round_index)

    def populate_table(self, table, category, round_index):
        headers = [
            "Rank",
            "Team ID", "Team Name", "Organization",
            "Payload", "Circuit", "Glide",
            "Loading", "Altitude", "Flight Score",
            "Takeoff", "Pilot", "Legal", "Landing", "Repl. Parts"
        ]

        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(self.teams))
        table.setSortingEnabled(False)
        table.verticalHeader().setVisible(False)

        font = QFont("Arial")
        font.setPointSize(12)

        best_Cdes, best_Tcarga, best_Tcircuit, best_Tglide = get_best_values_per_round(
            self.results, category, round_index
        )
        best_eff = best_Cdes / (best_Tcarga ** 0.5) if best_Tcarga > 0 else 1

        team_scores = []

        for row, team in enumerate(self.teams):
            tid = str(team["id"])
            name = team["name"]
            organization = team.get("organization", "")

            team_rounds = self.results.get(category, {}).get(tid, [])
            if round_index >= len(team_rounds):
                values = ["0.00"] * 6 + ["0.00"]
                total = 0.0
            else:
                entry = team_rounds[round_index]
                inputs = entry.get("inputs", {})

                Cdes = inputs.get("Unloaded Payload", 0)
                Csol = inputs.get("Requested Payload", 1)
                Tcircuit = inputs.get("Time Circuit", 1)
                Tglide = inputs.get("Time Glide", 0)
                A60s = inputs.get("Altitude", 0)
                Tcarga = inputs.get("Loading Time", 1)

                weights = {
                    "Academic": {
                        "payload": 150,
                        "circuit": 150,
                        "glide": 100,
                        "altitude": 100,
                        "loading": 100
                    },
                    "Clubs": {
                        "payload": 200,
                        "circuit": 200,
                        "glide": 150,
                        "altitude": 150,
                        "loading": 100
                    }
                }
                w = weights.get(category, weights["Academic"])

                payload_score = (
                    w["payload"] * (Cdes / best_Cdes) * (Cdes / Csol)
                    if Csol > 0 and best_Cdes > 0 else 0
                )
                circuit_score = (
                    w["circuit"] * (best_Tcircuit / Tcircuit)
                    if Tcircuit > 0 and best_Tcircuit > 0 else 0
                )
                glide_score = (
                    w["glide"] * (Tglide / best_Tglide)
                    if best_Tglide > 0 else 0
                )
                loading_score = (
                    w["loading"] * ((Cdes / (Tcarga ** 0.5)) / best_eff)
                    if Tcarga > 0 and best_eff > 0 else 0
                )

                if category == "Academic":
                    altitude_score = 4.3636e-6 * (A60s ** 4) - 0.001215 * (A60s ** 3) + 0.095732 * (A60s ** 2) - 0.86741 * A60s
                    altitude_score = math.ceil(altitude_score * 10) / 10
                    altitude_score = min(altitude_score, 100)
                elif category == "Clubs":
                    altitude_score = 6.5455e-6 * (A60s ** 4) - 0.001822 * (A60s ** 3) + 0.1436 * (A60s ** 2) - 1.3011 * A60s
                    altitude_score = math.ceil(altitude_score * 10) / 10
                    altitude_score = min(altitude_score, 150)
                else:
                    altitude_score = 0

                total = compute_round_score(
                    inputs, category,
                    best_unloaded_payload=best_Cdes,
                    best_loading_time=best_Tcarga,
                    best_circuit_time=best_Tcircuit,
                    best_glide_time=best_Tglide
                )

                values = [
                    f"{payload_score:.2f}",
                    f"{circuit_score:.2f}",
                    f"{glide_score:.2f}",
                    f"{loading_score:.2f}",
                    f"{altitude_score:.2f}",
                    f"{total:.2f}",
                ]

                takeoff = inputs.get("Takeoff Distance", "-")
                if isinstance(takeoff, (int, float)):
                    takeoff = str(int(round(takeoff)))
                pilot = "Team" if inputs.get("Pilot") else "External"
                legal = "✔️" if inputs.get("Legal Flight") else "❌"
                landing = "✔️" if inputs.get("Good Landing") else "❌"
                replacements = "No" if inputs.get("Replacement Parts") else "Yes"

                values += [str(takeoff), pilot, legal, landing, replacements]

            team_scores.append({
                "row": row,
                "total": float(values[5]) if len(values) > 5 else 0.0,
            })

            item_tid = NumericTableWidgetItem(tid, int(tid))
            item_tid.setFont(font)
            table.setItem(row, 1, item_tid)

            item_name = QTableWidgetItem(str(name))
            item_name.setFont(font)
            table.setItem(row, 2, item_name)

            item_org = QTableWidgetItem(str(organization))
            item_org.setFont(font)
            table.setItem(row, 3, item_org)

            for col, val in enumerate(values, start=4):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFont(font)

                if col == 9:  # Flight Score (total)
                    bold_font = QFont("Arial")
                    bold_font.setPointSize(12)
                    bold_font.setWeight(QFont.Weight.Bold)
                    item.setFont(bold_font)
                    item.setForeground(QColor("#228B22"))  # Forest Green
                elif col >= 10:
                    item.setForeground(QColor("#888888"))

                table.setItem(row, col, item)

        # Sort by total score descending and assign ranks with ties
        sorted_scores = sorted(team_scores, key=lambda x: x["total"], reverse=True)
        rank_map = {}
        current_rank = 1
        last_score = None
        for idx, entry in enumerate(sorted_scores):
            score = entry["total"]
            if last_score is not None and score < last_score:
                current_rank = idx + 1
            rank_map[entry["row"]] = current_rank
            last_score = score

        # Fill rank column
        for row in range(len(self.teams)):
            rank = rank_map[row]
            rank_item = NumericTableWidgetItem(str(rank), rank)
            rank_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            bold_font = QFont("Arial")
            bold_font.setPointSize(12)
            bold_font.setWeight(QFont.Weight.Bold)
            rank_item.setFont(bold_font)
            table.setItem(row, 0, rank_item)

        # Finally enable sorting on the table by Rank ascending
        table.setSortingEnabled(True)
        table.sortItems(0, Qt.SortOrder.AscendingOrder)

    def export_to_pdf(self):
        category = self.category_dropdown.currentText()
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results to PDF",
            f"{category}_round_results.pdf",
            "PDF Files (*.pdf)"
        )
        if not filename:
            return

        doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        for round_index in range(self.round_tabs.count()):
            elements.append(Paragraph(f"{category} - Round {round_index + 1}", styles["Title"]))
            elements.append(Spacer(1, 8))

            # Build data for the table
            headers = [
                "Rank", "Team ID", "Team Name", "Organization",
                "Payload", "Circuit", "Glide", "Loading", "Altitude", "Flight Score",
                "Takeoff", "Pilot", "Legal", "Landing", "Repl. Parts"
            ]
            data = [headers]

            tab = self.round_tabs.widget(round_index)
            table_widget = tab.layout().itemAt(0).widget()

            for row in range(table_widget.rowCount()):
                row_data = []
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            t = Table(data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t)
            elements.append(PageBreak())

        try:
            doc.build(elements)
            QMessageBox.information(self, "Success", f"PDF exported successfully to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export PDF:\n{e}")

    def export_to_excel(self):
        category = self.category_dropdown.currentText()
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results to Excel",
            f"{category}_round_results.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filename:
            return

        wb = Workbook()
        for round_index in range(self.round_tabs.count()):
            ws = wb.create_sheet(title=f"Round {round_index + 1}")
            tab = self.round_tabs.widget(round_index)
            table_widget = tab.layout().itemAt(0).widget()

            # Write headers
            for col in range(table_widget.columnCount()):
                header = table_widget.horizontalHeaderItem(col).text()
                ws.cell(row=1, column=col + 1, value=header)

            # Write data rows
            for row in range(table_widget.rowCount()):
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

            # Set column widths
            for col in range(table_widget.columnCount()):
                ws.column_dimensions[get_column_letter(col + 1)].width = 15

        # Remove default sheet if empty
        if "Sheet" in wb.sheetnames and not wb["Sheet"].max_row > 1:
            std = wb["Sheet"]
            wb.remove(std)

        try:
            wb.save(filename)
            QMessageBox.information(self, "Success", f"Excel file exported successfully to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel:\n{e}")
